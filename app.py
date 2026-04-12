from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_from_directory, send_file
import io
from flask_socketio import SocketIO, emit
import os
import threading
import logging
import glob
from datetime import datetime
import time

import pandas as pd
import random
import sys
import json
import mysql.connector
from dotenv import load_dotenv
from db_config import db_manager

# Cargar variables de entorno desde .env
load_dotenv()

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from scrapers.indeed_scraper import indeedScraper
from scrapers.linkedin_scraper import linkedinClass
from scrapers.linkedin_connection import linkedinConnections
from scrapers.computrabajo_scraper import ComputrabajoScraper
from scrapers.bumeran_scraper import BumeranScraper


from resume_parser import resumeParser, jobMatcher

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True

# WebSockets (async_mode=threading es compatible con el servidor dev de Flask)
socketio = SocketIO(app, async_mode='threading', cors_allowed_origins='*')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

for folder in ['uploads', 'results', 'logs', 'profiles']:
    os.makedirs(folder, exist_ok=True)

resume_parser = resumeParser()
job_matcher = jobMatcher()
job_status = {
    'indeed_scraping': {'status': 'idle', 'progress': 0, 'message': ''},
    'linkedin_scraping': {'status': 'idle', 'progress': 0, 'message': ''},
    'computrabajo_scraping': {'status': 'idle', 'progress': 0, 'message': ''},
    'bumeran_scraping': {'status': 'idle', 'progress': 0, 'message': ''},
    'universal_scraping': {'status': 'idle', 'progress': 0, 'message': '', 'scrapers': {
        'indeed': {'status': 'idle', 'progress': 0},
        'linkedin': {'status': 'idle', 'progress': 0},
        'computrabajo': {'status': 'idle', 'progress': 0},
        'bumeran': {'status': 'idle', 'progress': 0}
    }},
    'connection_requests': {'status': 'idle', 'progress': 0, 'message': ''}
}

# Simple in-memory task store for running quick test scrapers in background
test_tasks = {}

# Directory to persist task state so it is visible across Gunicorn workers
TASK_DIR = os.path.join('temp', 'test_tasks')
os.makedirs(TASK_DIR, exist_ok=True)

def _task_path(task_id: str) -> str:
    return os.path.join(TASK_DIR, f"{task_id}.json")

def save_task_to_disk(task_id: str):
    try:
        task = test_tasks.get(task_id)
        if task is None:
            return
        with open(_task_path(task_id), 'w', encoding='utf-8') as f:
            json.dump(task, f, ensure_ascii=False)
    except Exception:
        logger.exception(f"Failed to save task {task_id} to disk")

def load_task_from_disk(task_id: str):
    path = _task_path(task_id)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        logger.exception(f"Failed to load task {task_id} from disk")
        return None


def get_recent_job_files():
    csv_files = glob.glob(os.path.join('results', '*.csv'))
    files_with_time = [(f, os.path.getmtime(f)) for f in csv_files]
    files_with_time.sort(key=lambda x: x[1], reverse=True)
    return [os.path.basename(f[0]) for f in files_with_time[:10]]

def get_user_profile():
    """Load user profile from profiles/user_profiles.json"""
    try:
        profile_path = os.path.join('profiles', 'user_profiles.json')
        if os.path.exists(profile_path):
            with open(profile_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    except Exception as e:
        logger.error(f"Error loading user profile: {e}")
        return None

def save_user_profile(profile, filename):
    profile_path = os.path.join('profiles', 'user_profile.json')
    profile_data = profile.copy()
    profile_data['resume_filename'] = filename
    profile_data['upload_date'] = datetime.now().isoformat()
    
    try:
        with open(profile_path, 'w') as f:
            json.dump(profile_data, f, indent=2)
        return True
    except Exception as e:
        logger.error(f"Error saving profile: {e}")
        return False

def save_user_profile_for_role(profile, filename, role_type):
    profiles_dir = 'profiles'
    os.makedirs(profiles_dir, exist_ok=True)
    
    profiles_file = os.path.join(profiles_dir, 'user_profiles.json')
    
    if os.path.exists(profiles_file):
        with open(profiles_file, 'r') as f:
            all_profiles = json.load(f)
    else:
        all_profiles = {}
    
    profile_data = profile.copy()
    profile_data['resume_filename'] = filename
    profile_data['role_type'] = role_type
    profile_data['upload_date'] = datetime.now().isoformat()
    
    all_profiles[role_type] = profile_data
    
    try:
        with open(profiles_file, 'w') as f:
            json.dump(all_profiles, f, indent=2)
        logger.info(f"Saved profile for role: {role_type}")
        return True
    except Exception as e:
        logger.error(f"Error saving profile: {e}")
        return False

def get_all_user_profiles():
    profiles_file = os.path.join('profiles', 'user_profiles.json')
    if os.path.exists(profiles_file):
        try:
            with open(profiles_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error loading profiles: {e}")
    return {}

def get_user_profile_by_role(role_type):
    all_profiles = get_all_user_profiles()
    return all_profiles.get(role_type)

def delete_user_profile_by_role(role_type):
    profiles_file = os.path.join('profiles', 'user_profiles.json')
    
    if not os.path.exists(profiles_file):
        return False
    
    try:
        with open(profiles_file, 'r') as f:
            all_profiles = json.load(f)
        
        if role_type not in all_profiles:
            return False
        
        resume_filename = all_profiles[role_type].get('resume_filename')
        if resume_filename:
            resume_path = os.path.join('uploads', resume_filename)
            if os.path.exists(resume_path):
                try:
                    os.remove(resume_path)
                except OSError as e:
                    logger.warning(f"Could not delete resume file {resume_path}: {e}")
        
        del all_profiles[role_type]
        
        with open(profiles_file, 'w') as f:
            json.dump(all_profiles, f, indent=2)
        
        logger.info(f"Deleted profile for role: {role_type}")
        return True
        
    except Exception as e:
        logger.error(f"Error deleting profile: {e}")
        return False

def delete_user_profile():
    profile_path = os.path.join('profiles', 'user_profile.json')
    try:
        if os.path.exists(profile_path):
            os.remove(profile_path)
        
        upload_files = glob.glob(os.path.join('uploads', '*'))
        for file in upload_files:
            os.remove(file)
        
        return True
    except Exception as e:
        logger.error(f"Error deleting profile: {e}")
        return False


def get_user_profile():
    """Load user profile from profiles/user_profiles.json"""
    try:
        profile_path = os.path.join('profiles', 'user_profiles.json')
        if os.path.exists(profile_path):
            with open(profile_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    except Exception as e:
        logger.error(f"Error loading user profile: {e}")
        return None

def get_resume_text_from_profile():
    """Get resume text from any available profile (multi-resume system)"""
    try:
        # Intentar cargar todos los perfiles
        profile_path = os.path.join('profiles', 'user_profiles.json')
        if not os.path.exists(profile_path):
            return None
        
        with open(profile_path, 'r', encoding='utf-8') as f:
            all_profiles = json.load(f)
        
        # Si no hay perfiles, retornar None
        if not all_profiles:
            return None
        
        # Tomar el primer perfil disponible (cualquier rol)
        first_role = list(all_profiles.keys())[0]
        user_profile = all_profiles[first_role]
        
        resume_filename = user_profile.get('resume_filename')
        if not resume_filename:
            return None
        
        resume_path = os.path.join('uploads', resume_filename)
        if not os.path.exists(resume_path):
            return None
        
        return resume_parser.extract_text_from_file(resume_path)
        
    except Exception as e:
        logger.error(f"Error reading resume from profile: {e}")
        return None

def run_indeed_scraper_with_matching(credentials, searches, filters=None):
    try:
        job_status['indeed_scraping']['status'] = 'running'
        job_status['indeed_scraping']['progress'] = 5
        job_status['indeed_scraping']['message'] = 'Starting Indeed scraper...'
        
        indeed_email = credentials.get('indeed_email')
        indeed_password = credentials.get('indeed_password')
        
        all_profiles = get_user_profile()
        user_profile = None
        if all_profiles:
            # Clean up potential non-dict entries or empty keys
            valid_keys = [k for k in all_profiles.keys() if isinstance(all_profiles[k], dict)]
            if valid_keys:
                first_role = valid_keys[0]
                user_profile = all_profiles[first_role]
        
        user_skills = None
        if user_profile and 'resume_path' in user_profile:
            try:
                resume_path = user_profile['resume_path']
                if os.path.exists(resume_path):
                    with open(resume_path, 'r', encoding='utf-8', errors='ignore') as f:
                        resume_text = f.read()
                    user_skills = extract_skills_from_text(resume_text)
                    logger.info(f"Extracted {len(user_skills)} skills from resume")
            except Exception as e:
                logger.error(f"Error extracting skills: {e}")
        
        all_jobs = []
        scraper = None
        
        job_status['indeed_scraping']['progress'] = 10
        job_status['indeed_scraping']['message'] = 'Setting up scraper...'
        
        for i, search in enumerate(searches):
            position = search.get('position')
            location = search.get('location')
            
            search_progress_base = 20 + (i / len(searches)) * 60
            job_status['indeed_scraping']['progress'] = int(search_progress_base)
            job_status['indeed_scraping']['message'] = f"Searching for {position} in {location}... ({i+1}/{len(searches)})"
            
            if scraper:
                try:
                    scraper.close()
                except:
                    pass
            
            scraper = indeedScraper(indeed_email, indeed_password, user_skills=user_skills)
            
            try:
                if not scraper.login_with_google():
                    job_status['indeed_scraping']['message'] = f"Login failed. Skipping {position}."
                    continue
                
                jobs_per_search = 50
                
                if filters is None:
                    filters = {}
                clear_previous = (i == 0)
                jobs = scraper.search_jobs_with_filters(position, location, filters, clear_previous=clear_previous, max_jobs=jobs_per_search)
                
                if jobs:
                    new_jobs_count = len(jobs) - (len(all_jobs) - jobs_per_search) if i > 0 else len(jobs)
                    all_jobs = jobs
                    job_status['indeed_scraping']['message'] = f"Found {new_jobs_count} new jobs for {position}, Total: {len(all_jobs)}"
                else:
                    job_status['indeed_scraping']['message'] = f"No jobs found for {position}, Total: {len(all_jobs)}"
                
                time.sleep(random.uniform(2, 4))
                job_status['indeed_scraping']['progress'] = int(search_progress_base + 30)
                
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                logger.error(f"Error during search: {e}")
                logger.error(f"Full traceback:\n{error_trace}")
                job_status['indeed_scraping']['message'] = f"Error: {str(e)[:100]}"
                continue
        
            if all_jobs:
                job_status['indeed_scraping']['message'] = 'Processing results...'
                job_status['indeed_scraping']['progress'] = 85
                
                df = pd.DataFrame(all_jobs)
                
                if user_profile:
                    job_status['indeed_scraping']['message'] = 'Calculating matches...'
                    job_status['indeed_scraping']['progress'] = 90
                    
                    description_columns = ['Description', 'Job_Description', 'description', 'job_description']
                    description_col = None
                    
                    for col in description_columns:
                        if col in df.columns:
                            description_col = col
                            break
                    
                    if description_col:
                        df = job_matcher.process_job_dataframe(df, user_profile, description_col)
                        job_status['indeed_scraping']['message'] = 'Matching done'
                    else:
                        df['overall_match'] = 0.0
                        df['skill_match'] = 0.0
                        df['experience_match'] = 0.0
                        df['text_similarity'] = 0.0
                        df['matched_skills'] = [[] for _ in range(len(df))]
                
                job_status['indeed_scraping']['message'] = 'Saving...'
                job_status['indeed_scraping']['progress'] = 95
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename_suffix = "_with_matches" if user_profile else ""
                indeed_filename = f"indeed_jobs_all{filename_suffix}_{timestamp}.csv"
                
                file_path = os.path.join('results', indeed_filename)
                df.to_csv(file_path, index=False)
                
                job_status['indeed_scraping']['message'] = f"Saved {len(all_jobs)} jobs"
            else:
                job_status['indeed_scraping']['message'] = 'No jobs found'
        
        if scraper:
            try:
                scraper.close()
            except:
                pass
        
        job_status['indeed_scraping']['status'] = 'completed'
        job_status['indeed_scraping']['progress'] = 100
        
    except Exception as e:
        logger.error(f"Error in Indeed scraper: {str(e)}")
        job_status['indeed_scraping']['status'] = 'failed'
        job_status['indeed_scraping']['message'] = f"Error: {str(e)}"
        job_status['indeed_scraping']['progress'] = 0
        
        if scraper:
            try:
                scraper.close()
            except:
                pass

def run_linkedin_scraper_with_matching(credentials, searches, filters=None):
    if filters is None:
        filters = {}
    
    scraper = None
    
    try:
        job_status['linkedin_scraping']['status'] = 'running'
        job_status['linkedin_scraping']['progress'] = 5
        job_status['linkedin_scraping']['message'] = 'Starting LinkedIn scraper...'
        
        linkedin_token = credentials.get('linkedin_token')
        linkedin_email = credentials.get('linkedin_email')
        linkedin_password = credentials.get('linkedin_password')
        
        all_profiles = get_user_profile()
        user_profile = None
        if all_profiles:
            valid_keys = [k for k in all_profiles.keys() if isinstance(all_profiles[k], dict)]
            if valid_keys:
                first_role = valid_keys[0]
                user_profile = all_profiles[first_role]
        
        job_status['linkedin_scraping']['message'] = 'Initializing scraper...'
        try:
            # Try email/password first, fallback to token
            if linkedin_email and linkedin_password:
                scraper = linkedinClass(email=linkedin_email, password=linkedin_password)
            elif linkedin_token:
                scraper = linkedinClass(li_at_token=linkedin_token)
            else:
                job_status['linkedin_scraping']['status'] = 'failed'
                job_status['linkedin_scraping']['message'] = 'No LinkedIn credentials provided'
                job_status['linkedin_scraping']['progress'] = 0
                return
        except Exception as init_error:
            logger.error(f"Failed to initialize scraper: {init_error}")
            job_status['linkedin_scraping']['status'] = 'failed'
            job_status['linkedin_scraping']['message'] = f'Failed to initialize scraper: {str(init_error)}'
            job_status['linkedin_scraping']['progress'] = 0
            return
        
        job_status['linkedin_scraping']['message'] = 'Logging in to LinkedIn...'
        job_status['linkedin_scraping']['progress'] = 10
        
        login_success = False
        max_login_attempts = 2
        
        for attempt in range(max_login_attempts):
            try:
                # Use email/password login if available, otherwise use cookie
                if linkedin_email and linkedin_password:
                    login_success = scraper.login_with_credentials()
                else:
                    login_success = scraper.login_with_cookie()
                    
                if login_success:
                    logger.info("Login successful")
                    break
                else:
                    logger.warning(f"Login attempt {attempt + 1} failed")
                    time.sleep(5)
                    
                    if attempt < max_login_attempts - 1:
                        try:
                            if scraper and scraper.driver:
                                scraper.driver.quit()
                        except:
                            pass
                        time.sleep(2)
                        scraper = linkedinClass(linkedin_token)
                        
            except Exception as login_error:
                logger.error(f"Login error: {login_error}")
                if attempt == max_login_attempts - 1:
                    job_status['linkedin_scraping']['status'] = 'failed'
                    job_status['linkedin_scraping']['message'] = f'Login failed: {str(login_error)}'
                    job_status['linkedin_scraping']['progress'] = 0
                    return
        
        if not login_success:
            job_status['linkedin_scraping']['status'] = 'failed'
            job_status['linkedin_scraping']['message'] = 'Failed to login to LinkedIn'
            job_status['linkedin_scraping']['progress'] = 0
            return
            
        job_status['linkedin_scraping']['progress'] = 20
        job_status['linkedin_scraping']['message'] = 'Login successful, starting job searches...'
        
        all_jobs = []
        max_jobs_per_search = 50
        
        for i, search in enumerate(searches):
            position = search.get('position')
            location = search.get('location')
            
            search_progress_base = 20 + (i / len(searches)) * 60
            job_status['linkedin_scraping']['progress'] = int(search_progress_base)
            job_status['linkedin_scraping']['message'] = f"Searching for {position} in {location}... ({i+1}/{len(searches)})"
            
            try:
                jobs = scraper.search_jobs(
                    keyword=position, 
                    location=location, 
                    filters=filters,
                    max_jobs_per_search=max_jobs_per_search
                )
                
                if jobs:
                    all_jobs.extend(jobs)
                    job_status['linkedin_scraping']['message'] = f"Found {len(jobs)} jobs for {position}, Total: {len(all_jobs)}"
                    logger.info(f"Found {len(jobs)} jobs")
                else:
                    job_status['linkedin_scraping']['message'] = f"No jobs found for {position}, Total: {len(all_jobs)}"
                
                job_status['linkedin_scraping']['progress'] = int(search_progress_base + 30)
                
                if i < len(searches) - 1:
                    delay = random.uniform(3, 5)
                    time.sleep(delay)
                    
            except Exception as search_error:
                logger.error(f"Error searching: {search_error}")
                job_status['linkedin_scraping']['message'] = f"Error: {str(search_error)}"
                continue
        
        if all_jobs:
            job_status['linkedin_scraping']['message'] = 'Processing results...'
            job_status['linkedin_scraping']['progress'] = 85
            
            df = pd.DataFrame(all_jobs)
            
            if user_profile:
                job_status['linkedin_scraping']['message'] = 'Calculating matches...'
                job_status['linkedin_scraping']['progress'] = 90
                
                description_columns = ['Job_Description', 'Description', 'job_description', 'description']
                description_col = None
                
                for col in description_columns:
                    if col in df.columns:
                        description_col = col
                        break
                
                if description_col:
                    try:
                        df = job_matcher.process_job_dataframe(df, user_profile, description_col)
                        job_status['linkedin_scraping']['message'] = 'Matching done'
                    except Exception as match_error:
                        logger.error(f"Error in matching: {match_error}")
                else:
                    df['overall_match'] = 0.0
                    df['skill_match'] = 0.0
                    df['experience_match'] = 0.0
                    df['text_similarity'] = 0.0
                    df['matched_skills'] = [[] for _ in range(len(df))]
            
            job_status['linkedin_scraping']['message'] = 'Saving...'
            job_status['linkedin_scraping']['progress'] = 95
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_suffix = "_with_matches" if user_profile else ""
            linkedin_filename = f"linkedin_jobs_all{filename_suffix}_{timestamp}.csv"
            
            os.makedirs('results', exist_ok=True)
            
            file_path = os.path.join('results', linkedin_filename)
            
            try:
                df.to_csv(file_path, index=False)
                job_status['linkedin_scraping']['message'] = f"Saved {len(all_jobs)} jobs"
                logger.info(f"Saved {len(all_jobs)} jobs")
            except Exception as save_error:
                logger.error(f"Error saving: {save_error}")
                job_status['linkedin_scraping']['message'] = f"Error: {str(save_error)}"
                job_status['linkedin_scraping']['status'] = 'failed'
                return
                
        else:
            job_status['linkedin_scraping']['message'] = 'No jobs found'
        
        job_status['linkedin_scraping']['status'] = 'completed'
        job_status['linkedin_scraping']['progress'] = 100
        
    except Exception as e:
        logger.error(f"Error in LinkedIn scraper: {str(e)}")
        job_status['linkedin_scraping']['status'] = 'failed'
        job_status['linkedin_scraping']['message'] = f"Error: {str(e)}"
        job_status['linkedin_scraping']['progress'] = 0
        
    finally:
        if scraper:
            try:
                if hasattr(scraper, 'driver') and scraper.driver:
                    scraper.driver.quit()
            except Exception as cleanup_error:
                logger.error(f"Cleanup error: {cleanup_error}")

def run_connection_requests(credentials, company_list, message_template, search_mode='companies', search_term=''):
    try:
        job_status['connection_requests']['status'] = 'running'
        job_status['connection_requests']['progress'] = 5
        job_status['connection_requests']['message'] = 'Starting LinkedIn connection bot...'
        
        linkedin_email = credentials.get('linkedin_email')
        linkedin_password = credentials.get('linkedin_password')
        linkedin_token = credentials.get('linkedin_token')
        total_connections_sent = 0
        
        # Determine items to process based on mode
        items_to_process = []
        if search_mode == 'general':
            items_to_process = [search_term]
        else:
            items_to_process = company_list

        # Process each item with a fresh browser instance
        for index, item in enumerate(items_to_process):
            if job_status['connection_requests'].get('stop_requested'):
                job_status['connection_requests']['message'] = 'Process stopped by user.'
                job_status['connection_requests']['status'] = 'stopped'
                break

            bot = None
            
            search_progress_base = 20 + (index / len(items_to_process)) * 70
            job_status['connection_requests']['progress'] = int(search_progress_base)
            
            if search_mode == 'general':
                job_status['connection_requests']['message'] = f"Searching for: {item}..."
            else:
                job_status['connection_requests']['message'] = f"Processing connection requests for {item}..."
            
            try:
                bot = linkedinConnections(email=linkedin_email, password=linkedin_password, li_at_token=linkedin_token)
                
                login_success = False
                if linkedin_email and linkedin_password:
                    if bot.login_with_credentials():
                        login_success = True
                
                if not login_success and linkedin_token:
                    if bot.login_with_cookie():
                        login_success = True
                        
                if not login_success:
                    job_status['connection_requests']['message'] = f"Login failed. Stopping process."
                    logger.error("Login failed. Aborting remaining items.")
                    break
                    
                # Process item
                success = False
                if search_mode == 'general':
                    success = bot.process_general_search(
                        search_term=item,
                        message=message_template,
                        max_connections=25,
                        max_pages=5,
                        stop_check_callback=lambda: job_status['connection_requests'].get('stop_requested', False)
                    )
                else:
                    success = bot.process_company(
                        company=item, 
                        message=message_template,
                        max_connections=25,
                        max_pages=5,
                        stop_check_callback=lambda: job_status['connection_requests'].get('stop_requested', False)
                    )
                
                if job_status['connection_requests'].get('stop_requested'):
                    job_status['connection_requests']['message'] = 'Process stopped by user.'
                    job_status['connection_requests']['status'] = 'stopped'
                    break
                
                if success:
                    job_status['connection_requests']['message'] = f"Successfully processed {item}"
                    total_connections_sent += bot.total_connections_sent
                else:
                    job_status['connection_requests']['message'] = f"Failed to process {item}"
                    
            except Exception as e:
                logger.error(f"Unexpected error processing {item}: {str(e)}")
                job_status['connection_requests']['message'] = f"Error with {item}: {str(e)}"
            finally:
                # Always clean up the browser before moving to next item
                if bot:
                    bot.cleanup()
                    bot = None
                    
                # Add delay between items
                if index < len(company_list) - 1:  # Don't delay after the last company
                    delay = random.uniform(5, 10)
                    time.sleep(delay)
            
            # Update progress
            job_status['connection_requests']['progress'] = int(search_progress_base + 70/len(company_list))
        
        # Mark task as complete
        job_status['connection_requests']['status'] = 'completed'
        job_status['connection_requests']['progress'] = 100
        job_status['connection_requests']['message'] = f"Sent a total of {total_connections_sent} connection requests"
        
    except Exception as e:
        logger.error(f"Error in connection requests: {str(e)}")
        job_status['connection_requests']['status'] = 'failed'
        job_status['connection_requests']['message'] = f"Error: {str(e)}"
        job_status['connection_requests']['progress'] = 0

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scraper')
def scraper():
    user_profile = get_user_profile()
    profile_available = user_profile is not None
    return render_template('scraper.html', profile_available=profile_available)

@app.route('/results')
def results():
    job_files = get_recent_job_files()
    return render_template('results.html', job_files=job_files)


@app.route('/profile')
def profile():
    return render_template('profile.html')
import mysql.connector
from db_config import db_manager

@app.route('/view/<filename>')
def view_file(filename):
    try:
        file_path = os.path.join('results', filename)
        if os.path.exists(file_path) and filename.endswith('.csv'):
            # SOLUCIÓN: latin-1 puede decodificar cualquier byte (0-255) sin errores
            df = pd.read_csv(file_path, encoding='latin-1')
            
            has_matches = 'overall_match' in df.columns
            
            if has_matches:
                df = df.sort_values('overall_match', ascending=False)
            
            # Select columns to display (excluding match scores)
            # Define column mapping (English -> Spanish)
            column_mapping = {
                'Title': 'Título',
                'Company': 'Empresa',
                'Location': 'Ubicación',
                'Salary': 'Salario',
                'Job_Link': 'Enlace',
                'URL': 'URL',
                'Apply_URL': 'Apply_URL',
                'Experience_Category': 'Categoría de Experiencia'
            }

            # Rename columns in the main dataframe if they exist
            df = df.rename(columns=column_mapping)

            # Select columns to display (using Spanish names)
            display_columns = ['Título', 'Empresa', 'Ubicación', 'Enlace', 'URL', 'Apply_URL']
            
            display_columns = [col for col in display_columns if col in df.columns]
            
            display_df = df[display_columns].copy()
            
            # Make job links clickable
            if 'Enlace' in display_df.columns:
                display_df['Enlace'] = display_df['Enlace'].apply(
                    lambda x: f'<a href="{x}" target="_blank" class="btn btn-sm btn-primary">Postular</a>' 
                    if pd.notna(x) and str(x).startswith('http') else 'N/A'
                )
            
            table_html = display_df.head(100).to_html(
                classes='table table-striped table-hover table-sm', 
                table_id='jobTable',
                escape=False,
                index=False
            )
            
            return render_template('view_file.html', 
                                 filename=filename, 
                                 table_html=table_html,
                                 has_matches=has_matches,
                                 total_jobs=len(df),
                                 display_columns=display_columns)
        else:
            flash('File not found or not a CSV file', 'danger')
            return redirect(url_for('results'))
    except Exception as e:
        logger.error(f"Error viewing file: {str(e)}")
        flash(f'Error viewing file: {str(e)}', 'danger')
        return redirect(url_for('results'))

@app.route('/api/save_applications', methods=['POST'])
def api_save_applications():
    try:
        data = request.json
        filename = data.get('filename')
        applied_jobs = data.get('applied_jobs')
        total_jobs = data.get('total_jobs', 0)
        
        saved_ids = []
        companies = set()
        
        for job_data in applied_jobs.values():
            application_id = db_manager.save_application(job_data)
            if application_id:
                saved_ids.append(application_id)
                companies.add(job_data['company'])
        
        # Save file processing info
        db_manager.save_processed_file(filename, total_jobs, len(saved_ids))
        
        # Store in session for connection requests
        session['applied_companies'] = list(companies)
        
        return jsonify({
            'success': True,
            'saved': len(saved_ids),
            'companies': list(companies)
        })
        
    except Exception as e:
        logger.error(f"Error saving applications: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/exclude_company_from_networking', methods=['POST'])
def exclude_company_from_networking():
    try:
        company_name = request.form.get('company_name')
        if not company_name:
            return jsonify({'success': False, 'error': 'Company name required'}), 400
            
        # Remove from session if present
        if 'applied_companies' in session:
            if company_name in session['applied_companies']:
                session['applied_companies'].remove(company_name)
                session.modified = True
                
        if 'companies' in session:
            if company_name in session['companies']:
                session['companies'].remove(company_name)
                session.modified = True
        
        # Add to database exclusion list
        success = db_manager.exclude_company(company_name)
        
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Database error'}), 500
            
    except Exception as e:
        logger.error(f"Error excluding company: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/connections')
def connections():
    job_files = get_recent_job_files()
    
    auto_start = request.args.get('auto_start') == 'true'
    
    companies = session.get('applied_companies', [])
    
    if not companies:
        db_companies = db_manager.get_companies_for_connections()
        companies = [c['company'] for c in db_companies]
    
    return render_template('connections.html', 
                         job_files=job_files, 
                         companies=companies,
                         empresas=session.get('companies', companies),
                         auto_start=auto_start)

@app.route('/tracker')
def tracker():
    db_manager.create_job_tracker_table()
    return render_template('job_tracker.html')

@app.route('/api/tracker/jobs')
def api_get_tracked_jobs():
    try:
        jobs = db_manager.get_all_tracked_jobs()
        return jsonify({
            'success': True,
            'jobs': jobs,
            'count': len(jobs)
        })
    except Exception as e:
        logger.error(f"Error getting tracked jobs: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tracker/job', methods=['POST'])
def api_add_tracked_job():
    try:
        job_data = request.json
        
        if not job_data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        if not job_data.get('title') or not job_data.get('company'):
            return jsonify({'success': False, 'error': 'Title and Company are required'}), 400
        
        logger.info(f"Adding job to tracker: {job_data.get('title')} at {job_data.get('company')}")
        
        result = db_manager.add_job_to_tracker(job_data)
        
        if isinstance(result, dict) and result.get('exists'):
            logger.info(f"Job already exists with ID: {result['job_id']}")
            return jsonify({
                'success': False,
                'exists': True,
                'job_id': result['job_id'],
                'message': 'Job already exists in tracker'
            }), 409
        
        if result:
            logger.info(f"Successfully added job with ID: {result}")
            return jsonify({
                'success': True,
                'job_id': result,
                'message': 'Job added to tracker'
            })
        else:
            logger.error("Database returned None for job_id")
            return jsonify({'success': False, 'error': 'Database error: Failed to add job'}), 500
            
    except Exception as e:
        logger.error(f"Error adding job to tracker: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/api/tracker/job/<int:job_id>')
def api_get_tracked_job(job_id):
    try:
        job = db_manager.get_tracked_job_by_id(job_id)
        
        if job:
            return jsonify({
                'success': True,
                'job': job
            })
        else:
            return jsonify({'success': False, 'error': 'Job not found'}), 404
            
    except Exception as e:
        logger.error(f"Error getting tracked job: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tracker/job/<int:job_id>/status', methods=['PUT'])
def api_update_job_status(job_id):
    try:
        data = request.json
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'success': False, 'error': 'Status is required'}), 400
        
        success = db_manager.update_job_status(job_id, new_status)
        
        if success:
            return jsonify({
                'success': True,
                'message': f'Job status updated to {new_status}'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to update status'}), 500
            
    except Exception as e:
        logger.error(f"Error updating job status: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tracker/job/<int:job_id>', methods=['DELETE'])
def api_delete_tracked_job(job_id):
    try:
        success = db_manager.delete_tracked_job(job_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Job deleted successfully'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to delete job'}), 500
            
    except Exception as e:
        logger.error(f"Error deleting tracked job: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/get_applications')
def api_get_applications():
    try:
        filters = {
            'status': request.args.get('status'),
            'company': request.args.get('company')
        }
        
        applications = db_manager.get_applications(filters)
        
        return jsonify({
            'success': True,
            'applications': applications,
            'count': len(applications)
        })
        
    except Exception as e:
        logger.error(f"Error getting applications: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/view_profile')
def view_profile():
    user_profile = get_user_profile()
    if not user_profile:
        flash('No profile found. Please upload a resume first.', 'warning')
        return redirect(url_for('profile'))
    
    return render_template('view_profile.html', 
                         profile=user_profile, 
                         resume_filename=user_profile.get('resume_filename'))

@app.route('/upload_resume', methods=['POST'])
def upload_resume():
    if 'resume' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('profile'))
    
    file = request.files['resume']
    role_type = request.form.get('role_type')
    additional_skills = request.form.get('additional_skills', '')
    
    if not role_type:
        flash('Please select a role type', 'danger')
        return redirect(url_for('profile'))
    
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('profile'))
    
    if file:
        uploads_dir = 'uploads'
        os.makedirs(uploads_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"resume_{role_type}_{timestamp}_{file.filename}"
        filepath = os.path.join(uploads_dir, filename)
        file.save(filepath)
        
        try:
            profile = resume_parser.parse_resume(filepath)
            
            if profile:
                if additional_skills:
                    extra_skills = [s.strip() for s in additional_skills.split(',')]
                    profile['skills'].extend(extra_skills)
                    profile['skills'] = list(set(profile['skills']))
                
                if save_user_profile_for_role(profile, filename, role_type):
                    role_name = role_type.replace('_', ' ').title()
                    flash(f'✅ {role_name} resume uploaded successfully!', 'success')
                    return redirect(url_for('profile'))
                else:
                    flash('Error saving profile', 'danger')
            else:
                flash('Error parsing resume. Please check the file format.', 'danger')
                
        except Exception as e:
            logger.error(f"Error parsing resume: {str(e)}")
            flash(f'Error parsing resume: {str(e)}', 'danger')
    
    return redirect(url_for('profile'))

@app.route('/delete_profile')
def delete_profile():
    if delete_user_profile():
        flash('Profile deleted successfully', 'success')
    else:
        flash('Error deleting profile', 'danger')
    return redirect(url_for('profile'))

@app.route('/view_profile/<role_type>')
def view_profile_by_role(role_type):
    profile_data = get_user_profile_by_role(role_type)
    
    if not profile_data:
        flash(f'No profile found for {role_type.replace("_", " ").title()}', 'warning')
        return redirect(url_for('profile'))
    
    return render_template('view_profile.html', 
                         profile=profile_data, 
                         role_type=role_type,
                         resume_filename=profile_data.get('resume_filename'))

@app.route('/delete_profile/<role_type>')
def delete_profile_by_role(role_type):
    # Show confirmation page instead of deleting directly
    profile_data = get_user_profile_by_role(role_type)
    if profile_data:
        role_name = role_type.replace('_', ' ').title()
        return render_template('confirm_delete.html', role_type=role_type, role_name=role_name)
    else:
        flash('Profile not found', 'danger')
        return redirect(url_for('profile'))

@app.route('/delete_profile/<role_type>/confirm', methods=['POST'])
def confirm_delete_profile(role_type):
    if delete_user_profile_by_role(role_type):
        role_name = role_type.replace('_', ' ').title()
        flash(f'✅ {role_name} profile deleted successfully', 'success')
    else:
        flash('Error deleting profile', 'danger')
    return redirect(url_for('profile'))

@app.route('/api/all_profiles')
def api_all_profiles():
    profiles = get_all_user_profiles()
    return jsonify({
        'success': True,
        'profiles': profiles,
        'count': len(profiles)
    })

@app.route('/api/profile_summary')
def api_profile_summary():
    user_profile = get_user_profile()
    if not user_profile:
        return jsonify({'error': 'No profile found'}), 404
    
    return jsonify({
        'skills_count': len(user_profile.get('skills', [])),
        'experience_years': user_profile.get('experience_years', 0),
        'education_count': len(user_profile.get('education', [])),
        'upload_date': user_profile.get('upload_date')
    })

@app.route('/debug/directories')
def debug_directories():
    directories = {
        'uploads': os.path.exists('uploads'),
        'results': os.path.exists('results'),
        'logs': os.path.exists('logs'),
        'profiles': os.path.exists('profiles'),
        'temp': os.path.exists('temp'),
        'temp/resumes': os.path.exists('temp/resumes')
    }
    
    # Try to create directories if they don't exist
    for dir_path in ['uploads', 'results', 'logs', 'profiles', 'temp', 'temp/resumes']:
        try:
            os.makedirs(dir_path, exist_ok=True)
            directories[dir_path] = True
        except Exception as e:
            directories[dir_path] = f"Error: {str(e)}"
    
    return jsonify({
        'directories': directories,
        'current_working_dir': os.getcwd(),
        'message': 'Directory status checked and created if needed'
    })

@app.route('/status')
def status():
    return jsonify(job_status)

@app.route('/results/<filename>')
def download_file(filename):
    return send_from_directory('results', filename, as_attachment=True)


@app.route('/download_csv/<filename>')
def download_csv_file(filename):
    try:
        file_path = os.path.join('results', filename)
        if os.path.exists(file_path) and filename.endswith('.csv'):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash('File not found or not a CSV file', 'danger')
            return redirect(url_for('results'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'danger')
        return redirect(url_for('results'))

@app.route('/extract_companies/<filename>')
def extract_companies_from_file(filename):
    try:
        file_path = os.path.join('results', filename)
        if os.path.exists(file_path) and filename.endswith('.csv'):
            df = pd.read_csv(file_path, encoding='latin-1')
            
            # Extract companies
            companies = set()
            if 'Company' in df.columns:
                for company in df['Company'].dropna().unique():
                    # Clean company name - remove Inc., Corp., etc.
                    company_name = company.split(',')[0].split('Inc.')[0].split('Corp.')[0].strip()
                    if company_name and len(company_name) > 1:  # Ensure not empty
                        companies.add(company_name)
            
            # Filter out very common companies that might be too broad
            common_companies = ['amazon', 'google', 'microsoft', 'apple', 'facebook', 'meta']
            filtered_companies = [
                company for company in companies 
                if company.lower() not in common_companies
            ]
            
            # Limit to top 20 and sort alphabetically
            top_companies = sorted(filtered_companies)[:20]
            
            # Store in session
            session['companies'] = top_companies
            
            flash(f'Extracted {len(top_companies)} companies', 'success')
            return redirect(url_for('connections'))
        else:
            flash('File not found or not a CSV file', 'danger')
            return redirect(url_for('results'))
    except Exception as e:
        flash(f'Error extracting companies: {str(e)}', 'danger')
        return redirect(url_for('results'))


@app.route('/start_indeed_scraper', methods=['POST'])
def start_indeed_scraper():
    # Forzar recarga de .env para asegurar que tenemos las últimas credenciales
    load_dotenv(override=True)
    
    env_email = os.getenv('INDEED_EMAIL')
    env_password = os.getenv('INDEED_PASSWORD')
    
    logger.info(f"DEBUG: INDEED_EMAIL from env: {'Found' if env_email else 'Not Found'}")
    logger.info(f"DEBUG: INDEED_PASSWORD from env: {'Found' if env_password else 'Not Found'}")
    
    indeed_email = request.form.get('indeed_email') or env_email
    indeed_password = request.form.get('indeed_password') or env_password
    
    searches = []
    positions = request.form.getlist('position[]')
    locations = request.form.getlist('location[]')
    
    for i in range(len(positions)):
        if positions[i] and locations[i]:
            searches.append({
                'position': positions[i],
                'location': locations[i]
            })
    
    filters = {
        'salary_range': request.form.get('salary_range', ''),
        'remote_work': request.form.get('remote_work', ''),
        'experience_level': request.form.get('experience_level', ''),
        'education_level': request.form.get('education_level', ''),
        'date_posted': request.form.get('date_posted', '')
    }
    
    if not indeed_email or not indeed_password:
        flash('Please provide Indeed email and password', 'danger')
        return redirect(url_for('scraper'))
        
    if not searches:
        flash('Please provide at least one valid search', 'danger')
        return redirect(url_for('scraper'))
    
    job_status['indeed_scraping'] = {
        'status': 'idle', 
        'progress': 0, 
        'message': 'Starting...'
    }
    
    thread = threading.Thread(
        target=run_indeed_scraper_with_matching,
        args=({
            'indeed_email': indeed_email,
            'indeed_password': indeed_password
        }, searches, filters)
    )
    thread.daemon = True
    thread.start()
    
    flash('Indeed scraper started', 'success')
    return redirect(url_for('scraper'))

@app.route('/start_linkedin_scraper', methods=['POST'])
def start_linkedin_scraper():
    # Force reload of .env to get latest credentials
    load_dotenv(override=True)
    
    linkedin_token = request.form.get('linkedin_token')
    
    # If not provided in form, try to get from environment
    if not linkedin_token:
        linkedin_token = os.getenv('LI_AT_TOKEN')
    
    # Also read email/password from environment
    linkedin_email = os.getenv('LINKEDIN_EMAIL')
    linkedin_password = os.getenv('LINKEDIN_PASSWORD')
    
    searches = []
    positions = request.form.getlist('position[]')
    locations = request.form.getlist('location[]')
    
    for i in range(len(positions)):
        if positions[i] and locations[i]:
            searches.append({
                'position': positions[i],
                'location': locations[i]
            })
    
    filters = {
        'sort_by': request.form.get('sort_by', ''),
        'date_posted': request.form.get('date_posted', ''),
        'job_type': request.form.getlist('job_type[]'),
        'experience_level': request.form.get('experience_level', ''),
        'remote_work': request.form.get('remote_work', ''),
        'salary_range': request.form.get('salary_range', ''),
        'has_verifications': request.form.get('has_verifications', 'true') == 'true'
    }
    
    if not linkedin_token and not (linkedin_email and linkedin_password):
        flash('Please provide LinkedIn credentials (either token or email/password)', 'danger')
        return redirect(url_for('scraper'))
        
    if not searches:
        flash('Please provide at least one valid search', 'danger')
        return redirect(url_for('scraper'))
    
    job_status['linkedin_scraping'] = {
        'status': 'idle', 
        'progress': 0, 
        'message': 'Starting...'
    }
    
    thread = threading.Thread(
        target=run_linkedin_scraper_with_matching,
        args=({
            'linkedin_token': linkedin_token,
            'linkedin_email': linkedin_email,
            'linkedin_password': linkedin_password
        }, searches, filters)
    )
    thread.daemon = True
    thread.start()
    
    flash('LinkedIn scraper started', 'success')
    return redirect(url_for('scraper'))



@app.route('/start_computrabajo_scraper', methods=['POST'])
def start_computrabajo_scraper():
    """Start Computrabajo scraper"""
    # Force reload of .env to ensure we have the latest credentials
    load_dotenv(override=True)
    
    email = os.getenv('COMPUTRABAJO_EMAIL')
    password = os.getenv('COMPUTRABAJO_PASSWORD')
    
    searches = []
    positions = request.form.getlist('position[]')
    locations = request.form.getlist('location[]')
    
    for i in range(len(positions)):
        if positions[i]:
            searches.append({
                'position': positions[i],
                'location': locations[i] if i < len(locations) else ''
            })
    
    if not searches:
        flash('Please provide at least one valid search', 'danger')
        return redirect(url_for('scraper'))
    
    job_status['computrabajo_scraping'] = {
        'status': 'running', 
        'progress': 0, 
        'message': 'Starting Computrabajo scraper...'
    }
    
    def run_computrabajo_scraper():
        try:
            scraper = ComputrabajoScraper()
            all_jobs = []
            
            for i, search in enumerate(searches):
                position = search.get('position')
                location = search.get('location', '')
                
                job_status['computrabajo_scraping']['progress'] = int((i / len(searches)) * 80)
                job_status['computrabajo_scraping']['message'] = f"Searching for {position}..."
                
                # Pass credentials to search_jobs
                jobs = scraper.search_jobs(position, location, max_jobs=50, email=email, password=password)
                
                if jobs:
                    all_jobs.extend(jobs)
                    job_status['computrabajo_scraping']['message'] = f"Found {len(jobs)} jobs for {position}"
                
                time.sleep(random.uniform(2, 4))
            
            if all_jobs:
                job_status['computrabajo_scraping']['progress'] = 90
                job_status['computrabajo_scraping']['message'] = 'Saving results...'
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"computrabajo_jobs_{timestamp}.csv"
                scraper.save_to_csv(all_jobs, filename)
                
                job_status['computrabajo_scraping']['status'] = 'completed'
                job_status['computrabajo_scraping']['progress'] = 100
                job_status['computrabajo_scraping']['message'] = f'Saved {len(all_jobs)} jobs'
            else:
                job_status['computrabajo_scraping']['status'] = 'completed'
                job_status['computrabajo_scraping']['progress'] = 100
                job_status['computrabajo_scraping']['message'] = 'No jobs found'
            
            scraper.close()
            
        except Exception as e:
            logger.error(f"Error in Computrabajo scraper: {e}")
            job_status['computrabajo_scraping']['status'] = 'failed'
            job_status['computrabajo_scraping']['message'] = f'Error: {str(e)[:100]}'
    
    thread = threading.Thread(target=run_computrabajo_scraper)
    thread.daemon = True
    thread.start()
    
    flash('Computrabajo scraper started', 'success')
    return redirect(url_for('scraper'))


@app.route('/start_bumeran_scraper', methods=['POST'])
def start_bumeran_scraper():
    """Start Bumeran scraper"""
    searches = []
    positions = request.form.getlist('position[]')
    locations = request.form.getlist('location[]')
    
    for i in range(len(positions)):
        if positions[i]:
            searches.append({
                'position': positions[i],
                'location': locations[i] if i < len(locations) else ''
            })
    
    if not searches:
        flash('Please provide at least one valid search', 'danger')
        return redirect(url_for('scraper'))
    
    job_status['bumeran_scraping'] = {
        'status': 'running', 
        'progress': 0, 
        'message': 'Starting Bumeran scraper...'
    }
    
    def run_bumeran_scraper():
        try:
            # Get credentials
            email = os.getenv('BUMERAN_EMAIL')
            password = os.getenv('BUMERAN_PASSWORD')
            
            scraper = BumeranScraper()
            all_jobs = []
            
            for i, search in enumerate(searches):
                position = search.get('position')
                location = search.get('location', '')
                
                job_status['bumeran_scraping']['progress'] = int((i / len(searches)) * 80)
                job_status['bumeran_scraping']['message'] = f"Searching for {position}..."
                
                # Pass credentials to search_jobs
                jobs = scraper.search_jobs(position, location, max_jobs=50, email=email, password=password)
                
                if jobs:
                    all_jobs.extend(jobs)
                    job_status['bumeran_scraping']['message'] = f"Found {len(jobs)} jobs for {position}"
                
                time.sleep(random.uniform(2, 4))
            
            if all_jobs:
                job_status['bumeran_scraping']['progress'] = 90
                job_status['bumeran_scraping']['message'] = 'Saving results...'
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"bumeran_jobs_{timestamp}.csv"
                scraper.save_to_csv(all_jobs, filename)
                
                job_status['bumeran_scraping']['status'] = 'completed'
                job_status['bumeran_scraping']['progress'] = 100
                job_status['bumeran_scraping']['message'] = f'Saved {len(all_jobs)} jobs'
            else:
                job_status['bumeran_scraping']['status'] = 'completed'
                job_status['bumeran_scraping']['progress'] = 100
                job_status['bumeran_scraping']['message'] = 'No jobs found'
            
            scraper.close()
            
        except Exception as e:
            logger.error(f"Error in Bumeran scraper: {e}")
            job_status['bumeran_scraping']['status'] = 'failed'
            job_status['bumeran_scraping']['message'] = f'Error: {str(e)[:100]}'
    
    thread = threading.Thread(target=run_bumeran_scraper)
    thread.daemon = True
    thread.start()
    
    flash('Bumeran scraper started', 'success')
    return redirect(url_for('scraper'))


@app.route('/run_test_scraper', methods=['GET'])
def run_test_scraper():
    """Run a quick test for a specific scraper and return results as JSON.
    Example: /run_test_scraper?scraper=bumeran&q=analista&location=Buenos+Aires&max=5
    """
    scraper_name = (request.args.get('scraper') or 'bumeran').lower()
    position = request.args.get('q') or request.args.get('position') or ''
    location = request.args.get('location', '')
    try:
        max_jobs = int(request.args.get('max', 5))
    except ValueError:
        max_jobs = 5

    try:
        if scraper_name == 'bumeran':
            email = os.getenv('BUMERAN_EMAIL')
            password = os.getenv('BUMERAN_PASSWORD')
            scraper = BumeranScraper()
            jobs = scraper.search_jobs(position, location, max_jobs=max_jobs, email=email, password=password)

        elif scraper_name == 'computrabajo':
            scraper = ComputrabajoScraper()
            jobs = scraper.search_jobs(position, location, max_jobs_per_search=max_jobs)

        elif scraper_name == 'indeed':
            scraper = indeedScraper()
            jobs = scraper.search_jobs(position, location, max_jobs_per_search=max_jobs)

        elif scraper_name == 'linkedin':
            # Run LinkedIn scraper in background to avoid worker timeouts
            li_at = request.args.get('linkedin_token') or os.getenv('LINKEDIN_LI_AT') or os.getenv('LI_AT') or os.getenv('LINKEDIN_TOKEN')
            if not li_at:
                return jsonify({'error': 'Please provide LinkedIn authentication token via ?linkedin_token=... or env LINKEDIN_LI_AT/LI_AT'}), 400

            import uuid

            task_id = str(uuid.uuid4())
            test_tasks[task_id] = {
                'status': 'queued',
                'progress': 0,
                'scraper': 'linkedin',
                'position': position,
                'location': location,
                'max': max_jobs,
                'result': None,
                'error': None,
                'started_at': None,
                'finished_at': None
            }
            # persist initial state
            save_task_to_disk(task_id)

            def run_linkedin_task(tid, token, position, location, max_jobs):
                test_tasks[tid]['status'] = 'running'
                test_tasks[tid]['started_at'] = datetime.now().isoformat()
                save_task_to_disk(tid)
                try:
                    scraper = linkedinClass(token)
                    try:
                        test_tasks[tid]['progress'] = 10
                        save_task_to_disk(tid)
                        logged = getattr(scraper, 'login_with_cookie', lambda: True)()
                    except Exception:
                        logged = False
                    test_tasks[tid]['progress'] = 30
                    save_task_to_disk(tid)
                    jobs = scraper.search_jobs(position, location, max_jobs_per_search=max_jobs)
                    # normalize jobs
                    simple_jobs = []
                    for j in (jobs or [])[:max_jobs]:
                        if isinstance(j, dict):
                            simple_jobs.append(j)
                        else:
                            try:
                                simple_jobs.append(j.__dict__)
                            except Exception:
                                simple_jobs.append(str(j))

                    test_tasks[tid]['result'] = simple_jobs
                    test_tasks[tid]['status'] = 'completed'
                    test_tasks[tid]['progress'] = 100
                    save_task_to_disk(tid)
                except Exception as e:
                    logger.exception(f"Error in background LinkedIn task {tid}: {e}")
                    test_tasks[tid]['error'] = str(e)[:1000]
                    test_tasks[tid]['status'] = 'failed'
                    test_tasks[tid]['progress'] = 0
                    save_task_to_disk(tid)
                finally:
                    test_tasks[tid]['finished_at'] = datetime.now().isoformat()
                    save_task_to_disk(tid)
                    try:
                        close_fn = getattr(scraper, 'close', None)
                        if callable(close_fn):
                            close_fn()
                    except Exception:
                        pass

            thread = threading.Thread(target=run_linkedin_task, args=(task_id, li_at, position, location, max_jobs))
            thread.daemon = True
            thread.start()

            return jsonify({'task_id': task_id, 'status': 'started'}), 202

        else:
            return jsonify({'error': 'Unknown scraper: ' + scraper_name}), 400

        # Ensure we close the scraper if it offers a close method
        try:
            close_fn = getattr(scraper, 'close', None)
            if callable(close_fn):
                close_fn()
        except Exception:
            pass

        # Truncate job entries to simple serializable structure if needed
        simple_jobs = []
        for j in (jobs or [])[:max_jobs]:
            if isinstance(j, dict):
                simple_jobs.append(j)
            else:
                # Best-effort mapping for objects
                try:
                    simple_jobs.append(j.__dict__)
                except Exception:
                    simple_jobs.append(str(j))

        return jsonify({'scraper': scraper_name, 'count': len(simple_jobs), 'jobs': simple_jobs})

    except Exception as e:
        logger.exception(f"Error running test scraper {scraper_name}: {e}")
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/test_task_status', methods=['GET'])
def test_task_status():
    """Get status/result for a background test task: /test_task_status?task_id=..."""
    task_id = request.args.get('task_id')
    if not task_id:
        return jsonify({'error': 'task_id is required'}), 400

    task = test_tasks.get(task_id) or load_task_from_disk(task_id)
    if not task:
        return jsonify({'error': 'task_id not found'}), 404

    # Return limited fields
    response = {
        'task_id': task_id,
        'status': task.get('status'),
        'progress': task.get('progress'),
        'started_at': task.get('started_at'),
        'finished_at': task.get('finished_at'),
        'error': task.get('error'),
        'count': len(task['result']) if task.get('result') else 0,
        'result': task.get('result') if task.get('status') == 'completed' else None
    }

    return jsonify(response)



@app.route('/start_universal_scraper', methods=['POST'])
def start_universal_scraper():
    """Start Universal Scraper - runs all 4 scrapers simultaneously"""
    position = request.form.get('position')
    location = request.form.get('location', '')
    
    if not position:
        flash('Por favor proporciona un puesto para buscar', 'danger')
        return redirect(url_for('scraper'))
    
    # Initialize universal scraper status
    job_status['universal_scraping'] = {
        'status': 'running',
        'progress': 0,
        'message': 'Iniciando búsqueda en Computrabajo y Bumeran...',
        'scrapers': {
            'indeed': {'status': 'idle', 'progress': 0},
            'linkedin': {'status': 'idle', 'progress': 0},
            'computrabajo': {'status': 'idle', 'progress': 0},
            'bumeran': {'status': 'idle', 'progress': 0}
        }
    }
    
    def run_universal_scraper():
        try:
            all_results = []
            completed_scrapers = 0
            total_scrapers = 4
            scraper_threads = []
            scraper_results = {
                'indeed': [],
                'linkedin': [],
                'computrabajo': [],
                'bumeran': []
            }
            
            # Get credentials from environment
            indeed_email = os.getenv('INDEED_EMAIL')
            indeed_password = os.getenv('INDEED_PASSWORD')
            linkedin_token = os.getenv('LINKEDIN_TOKEN')
            computrabajo_email = os.getenv('COMPUTRABAJO_EMAIL')
            computrabajo_password = os.getenv('COMPUTRABAJO_PASSWORD')
            bumeran_email = os.getenv('BUMERAN_EMAIL')
            bumeran_password = os.getenv('BUMERAN_PASSWORD')
            
            # Function to run Indeed scraper
            def run_indeed():
                # Skipping Indeed in universal scraper - use individual scraper instead
                logger.info("Skipping Indeed in universal scraper")
                job_status['universal_scraping']['scrapers']['indeed']['status'] = 'skipped'
                job_status['universal_scraping']['scrapers']['indeed']['message'] = 'Usar scraper individual'
            
            # Function to run LinkedIn scraper
            def run_linkedin():
                # Skipping LinkedIn in universal scraper - use individual scraper instead
                logger.info("Skipping LinkedIn in universal scraper")
                job_status['universal_scraping']['scrapers']['linkedin']['status'] = 'skipped'
                job_status['universal_scraping']['scrapers']['linkedin']['message'] = 'Usar scraper individual'
            
            # Function to run Computrabajo scraper
            def run_computrabajo():
                try:
                    job_status['universal_scraping']['scrapers']['computrabajo']['status'] = 'running'
                    job_status['universal_scraping']['message'] = 'Buscando en Computrabajo...'
                    
                    scraper = ComputrabajoScraper()
                    jobs = scraper.search_jobs(position, location, max_jobs=50, 
                                              email=computrabajo_email, password=computrabajo_password)
                    
                    if jobs:
                        for job in jobs:
                            job['Fuente'] = 'Computrabajo'
                        scraper_results['computrabajo'] = jobs
                    
                    scraper.close()
                    job_status['universal_scraping']['scrapers']['computrabajo']['status'] = 'completed'
                    job_status['universal_scraping']['scrapers']['computrabajo']['progress'] = 100
                except Exception as e:
                    logger.error(f"Error in Computrabajo scraper: {e}")
                    job_status['universal_scraping']['scrapers']['computrabajo']['status'] = 'failed'
            
            # Function to run Bumeran scraper
            def run_bumeran():
                try:
                    job_status['universal_scraping']['scrapers']['bumeran']['status'] = 'running'
                    job_status['universal_scraping']['message'] = 'Buscando en Bumeran...'
                    
                    scraper = BumeranScraper()
                    jobs = scraper.search_jobs(position, location, max_jobs=50,
                                             email=bumeran_email, password=bumeran_password)
                    
                    if jobs:
                        for job in jobs:
                            job['Fuente'] = 'Bumeran'
                        scraper_results['bumeran'] = jobs
                    
                    scraper.close()
                    job_status['universal_scraping']['scrapers']['bumeran']['status'] = 'completed'
                    job_status['universal_scraping']['scrapers']['bumeran']['progress'] = 100
                except Exception as e:
                    logger.error(f"Error in Bumeran scraper: {e}")
                    job_status['universal_scraping']['scrapers']['bumeran']['status'] = 'failed'
            
            # Start all scrapers in parallel
            indeed_thread = threading.Thread(target=run_indeed)
            linkedin_thread = threading.Thread(target=run_linkedin)
            computrabajo_thread = threading.Thread(target=run_computrabajo)
            bumeran_thread = threading.Thread(target=run_bumeran)
            
            indeed_thread.start()
            linkedin_thread.start()
            computrabajo_thread.start()
            bumeran_thread.start()
            
            # Wait for all to complete
            indeed_thread.join()
            linkedin_thread.join()
            computrabajo_thread.join()
            bumeran_thread.join()
            
            # Consolidate results
            job_status['universal_scraping']['message'] = 'Consolidando resultados...'
            job_status['universal_scraping']['progress'] = 90
            
            for source, jobs in scraper_results.items():
                all_results.extend(jobs)
            
            if all_results:
                # Create consolidated DataFrame
                df = pd.DataFrame(all_results)
                
                # Reorder columns to have Fuente first
                cols = df.columns.tolist()
                if 'Fuente' in cols:
                    cols.remove('Fuente')
                    cols = ['Fuente'] + cols
                    df = df[cols]
                
                # Save to CSV
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"universal_jobs_{timestamp}.csv"
                filepath = os.path.join('results', filename)
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                
                job_status['universal_scraping']['status'] = 'completed'
                job_status['universal_scraping']['progress'] = 100
                job_status['universal_scraping']['message'] = f'✅ Búsqueda completada: {len(all_results)} empleos encontrados en {len([s for s in scraper_results.values() if s])} plataformas'
            else:
                job_status['universal_scraping']['status'] = 'completed'
                job_status['universal_scraping']['progress'] = 100
                job_status['universal_scraping']['message'] = 'No se encontraron empleos'
                
        except Exception as e:
            logger.error(f"Error in Universal scraper: {e}")
            job_status['universal_scraping']['status'] = 'failed'
            job_status['universal_scraping']['message'] = f'Error: {str(e)[:100]}'
    
    thread = threading.Thread(target=run_universal_scraper)
    thread.daemon = True
    thread.start()
    
    flash('Búsqueda universal iniciada en todas las plataformas', 'success')
    return redirect(url_for('scraper'))



@app.route('/stop_connection_requests', methods=['POST'])
def stop_connection_requests():
    if 'connection_requests' in job_status:
        job_status['connection_requests']['stop_requested'] = True
        job_status['connection_requests']['message'] = 'Stopping process...'
        flash('Stopping connection requests...', 'warning')
    return redirect(url_for('connections'))

@app.route('/start_connection_requests', methods=['POST'])
def start_connection_requests():
    linkedin_email = request.form.get('linkedin_email')
    linkedin_password = request.form.get('linkedin_password')
    linkedin_token = request.form.get('linkedin_token')
    
    search_mode = request.form.get('search_mode', 'companies') # 'companies' or 'general'
    search_term = request.form.get('search_term', '')
    company_list = request.form.getlist('company_selection')
    message_template = request.form.get('message_template', '')
    
    if not (linkedin_email and linkedin_password) and not linkedin_token:
        flash('Please provide LinkedIn credentials (email/password) or token', 'danger')
        return redirect(url_for('connections'))
        
    if search_mode == 'companies' and not company_list:
        flash('Please select at least one company', 'danger')
        return redirect(url_for('connections'))
        
    if search_mode == 'general' and not search_term:
        flash('Please enter a search term', 'danger')
        return redirect(url_for('connections'))
    
    job_status['connection_requests'] = {
        'status': 'idle', 
        'progress': 0, 
        'message': 'Starting...',
        'stop_requested': False
    }
    
    thread = threading.Thread(
        target=run_connection_requests,
        args=({
            'linkedin_email': linkedin_email,
            'linkedin_password': linkedin_password,
            'linkedin_token': linkedin_token
        }, company_list, message_template, search_mode, search_term)
    )
    thread.daemon = True
    thread.start()
    
    flash('Connection requests started', 'success')
    return redirect(url_for('connections'))

@app.route('/download_resume/<filename>')
def download_resume(filename):
    try:
        # Check in cover_letters directory first
        cover_letters_dir = os.path.join('temp', 'cover_letters')
        if os.path.exists(os.path.join(cover_letters_dir, filename)):
            mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            return send_from_directory(
                cover_letters_dir, 
                filename, 
                as_attachment=True, 
                download_name=filename,
                mimetype=mimetype
            )
        
        # Fallback to resumes directory
        resume_dir = os.path.join('temp', 'resumes')
        mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        return send_from_directory(
            resume_dir, 
            filename, 
            as_attachment=True, 
            download_name=filename,
            mimetype=mimetype
        )
    except Exception as e:
        flash(f'Error downloading resume: {str(e)}', 'danger')
        return redirect(url_for('results'))

@app.route('/delete_csv/<filename>')
def delete_csv(filename):
    try:
        results_dir = 'results'
        filepath = os.path.join(results_dir, filename)
        
        if os.path.exists(filepath):
            os.remove(filepath)
            flash(f'File {filename} deleted successfully!', 'success')
        else:
            flash(f'File {filename} not found!', 'warning')
        
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Error deleting file: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/batch_cover_letters')
def batch_cover_letters():
    try:
        results_dir = 'results'
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)
        
        job_files = [f for f in os.listdir(results_dir) if f.endswith('.csv')]
        job_files.sort(reverse=True)
        
        return render_template('batch_cover_letters.html', job_files=job_files)
    except Exception as e:
        logger.error(f"Error loading batch cover letters page: {e}")
        flash('Error loading page', 'danger')
        return redirect(url_for('index'))

@app.route('/api/batch_cover_letters', methods=['POST'])
def api_batch_cover_letters():
    try:
        data = request.json
        filename = data.get('filename')
        num_jobs = int(data.get('num_jobs', 5))
        
        if not filename:
            return jsonify({'success': False, 'error': 'Filename is required'}), 400
        
        resume_text = get_resume_text_from_profile()
        if not resume_text:
            return jsonify({'success': False, 'error': 'Please upload your resume first'}), 400
        
        file_path = os.path.join('results', filename)
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        df = pd.read_csv(file_path, encoding='latin-1')
        
        description_col = None
        for col in ['Job_Description', 'Description', 'job_description', 'description']:
            if col in df.columns:
                description_col = col
                break
        
        if not description_col:
            return jsonify({'success': False, 'error': 'No job description column found'}), 400
        
        generator = CoverLetterGenerator()
        results = []
        
        jobs_to_process = df.head(num_jobs)
        
        for index, row in jobs_to_process.iterrows():
            try:
                job_title = row.get('Title', 'Unknown Position')
                company = row.get('Company', 'Unknown Company')
                job_description = str(row.get(description_col, ''))
                
                if not job_description or job_description == 'nan':
                    continue
                
                cover_letter = generator.generate_cover_letter(resume_text, job_description, job_title, company)
                saved_filename = generator.save_cover_letter(cover_letter, job_title, company)
                
                results.append({
                    'job_title': job_title,
                    'company': company,
                    'filename': saved_filename,
                    'status': 'success'
                })
                
                time.sleep(2)
                
            except Exception as e:
                logger.error(f"Error generating cover letter for job {index}: {e}")
                results.append({
                    'job_title': row.get('Title', 'Unknown'),
                    'company': row.get('Company', 'Unknown'),
                    'filename': None,
                    'status': 'failed',
                    'error': str(e)
                })
        
        successful = sum(1 for r in results if r['status'] == 'success')
        
        return jsonify({
            'success': True,
            'results': results,
            'total': len(results),
            'successful': successful
        })
        
    except Exception as e:
        logger.error(f"Error in batch cover letter generation: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/optimize_resume')
def optimize_resume_page():
    return render_template('optimize_resume.html')

@app.route('/optimize_resume', methods=['POST'])
def optimize_resume():
    try:
        job_title = request.form.get('job_title', '')
        company = request.form.get('company', '')
        job_description = request.form.get('job_description', '')
        resume_text = request.form.get('resume_text', '')
        
        if not job_description:
            flash('Job description is required', 'danger')
            return redirect(url_for('optimize_resume_page'))
        
        if not resume_text:
            resume_text = get_resume_text_from_profile()
            if not resume_text:
                flash('Please provide resume text or upload your resume first', 'warning')
                return redirect(url_for('optimize_resume_page'))
        
        optimizer = simpleResumeOptimizer()
        optimized_resume = optimizer.optimize_resume(resume_text, job_description, job_title, company)
        
        filename = optimizer.save_resume(optimized_resume, job_title, company)
        
        return render_template('optimize_resume.html', 
                             optimized_resume=optimized_resume,
                             filename=filename,
                             job_title=job_title,
                             company=company)
        
    except Exception as e:
        logger.error(f"Error optimizing resume: {e}")
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('optimize_resume_page'))

@app.route('/api/optimize_resume', methods=['POST'])
def api_optimize_resume():
    try:
        data = request.json
        job_title = data.get('job_title', '')
        company = data.get('company', '')
        job_description = data.get('job_description', '')
        
        if not job_description:
            return jsonify({'success': False, 'error': 'Job description is required'}), 400
        
        resume_text = get_resume_text_from_profile()
        if not resume_text:
            return jsonify({'success': False, 'error': 'Please upload your resume first'}), 400
        
        optimizer = simpleResumeOptimizer()
        optimized_resume = optimizer.optimize_resume(resume_text, job_description, job_title, company)
        
        filename = optimizer.save_resume(optimized_resume, job_title, company)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'resume_text': optimized_resume
        })
        
    except Exception as e:
        logger.error(f"Error optimizing resume: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/batch_optimize')
def batch_optimize_page():
    job_files = get_recent_job_files()
    return render_template('batch_optimize.html', job_files=job_files)

@app.route('/api/batch_optimize', methods=['POST'])
def api_batch_optimize():
    try:
        data = request.json
        filename = data.get('filename')
        num_jobs = int(data.get('num_jobs', 5))
        
        if not filename:
            return jsonify({'success': False, 'error': 'Filename is required'}), 400
        
        resume_text = get_resume_text_from_profile()
        if not resume_text:
            return jsonify({'success': False, 'error': 'Please upload your resume first'}), 400
        
        file_path = os.path.join('results', filename)
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        df = pd.read_csv(file_path, encoding='latin-1')
        
        description_col = None
        for col in ['Job_Description', 'Description', 'job_description', 'description']:
            if col in df.columns:
                description_col = col
                break
        
        if not description_col:
            return jsonify({'success': False, 'error': 'No job description column found'}), 400
        
        optimizer = simpleResumeOptimizer()
        results = []
        
        jobs_to_process = df.head(num_jobs)
        
        for index, row in jobs_to_process.iterrows():
            try:
                job_title = row.get('Title', 'Unknown Position')
                company = row.get('Company', 'Unknown Company')
                job_description = str(row.get(description_col, ''))
                
                if not job_description or job_description == 'nan':
                    continue
                
                optimized_resume = optimizer.optimize_resume(resume_text, job_description, job_title, company)
                saved_filename = optimizer.save_resume(optimized_resume, job_title, company)
                
                results.append({
                    'job_title': job_title,
                    'company': company,
                    'filename': saved_filename,
                    'status': 'success'
                })
                
                time.sleep(2)
                
            except Exception as e:
                logger.error(f"Error optimizing for job {index}: {e}")
                results.append({
                    'job_title': row.get('Title', 'Unknown'),
                    'company': row.get('Company', 'Unknown'),
                    'filename': None,
                    'status': 'failed',
                    'error': str(e)
                })
        
        return jsonify({
            'success': True,
            'results': results,
            'total': len(results),
            'successful': len([r for r in results if r['status'] == 'success'])
        })
        
    except Exception as e:
        logger.error(f"Error in batch optimization: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ─────────────────────────────────────────────────────────────
# NUEVAS RUTAS: Analytics, Tracker Stats, Error Handlers
# ─────────────────────────────────────────────────────────────

@app.route('/analytics')
def analytics():
    """Dashboard de analytics: métricas de postulaciones y archivos scrapeados."""
    from collections import defaultdict

    # ── Datos desde MySQL (tabla applications) ──
    stats = {
        'total_jobs': 0,
        'bookmarked': 0,
        'applying': 0,
        'applied': 0,
        'interviewing': 0,
        'negotiating': 0,
        'accepted': 0,
        'rejected': 0,
        'top_companies': [],
    }

    timeline_labels = []
    timeline_data   = []
    status_labels   = []
    status_data     = []
    csv_stats       = []

    try:
        all_jobs = db_manager.get_all_tracked_jobs()
        status_counts = defaultdict(int)
        company_counts = defaultdict(int)
        date_counts    = defaultdict(int)

        for job in all_jobs:
            st = job.get('status', 'unknown')
            status_counts[st] += 1
            company = job.get('company', 'Desconocida')
            company_counts[company] += 1
            date_applied = job.get('applied_date') or job.get('date_saved')
            if date_applied:
                try:
                    day = str(date_applied)[:10]
                    date_counts[day] += 1
                except Exception:
                    pass

        stats['total_jobs']    = len(all_jobs)
        stats['bookmarked']    = status_counts.get('bookmarked', 0)
        stats['applying']      = status_counts.get('applying', 0)
        stats['applied']       = status_counts.get('applied', 0)
        stats['interviewing']  = status_counts.get('interviewing', 0)
        stats['negotiating']   = status_counts.get('negotiating', 0)
        stats['accepted']      = status_counts.get('accepted', 0)
        stats['rejected']      = status_counts.get('rejected', 0)

        # Top companies
        stats['top_companies'] = [
            {'name': k, 'count': v}
            for k, v in sorted(company_counts.items(), key=lambda x: -x[1])[:10]
        ]

        # Timeline (últimos 30 días)
        sorted_dates = sorted(date_counts.keys())[-30:]
        timeline_labels = sorted_dates
        timeline_data   = [date_counts[d] for d in sorted_dates]

        # Status doughnut
        status_order = ['applied','interviewing','negotiating','accepted','bookmarked','applying','rejected']
        status_labels_map = {
            'applied': 'Postulados',
            'interviewing': 'Entrevistas',
            'negotiating': 'Negociando',
            'accepted': 'Aceptados',
            'bookmarked': 'Guardados',
            'applying': 'Aplicando',
            'rejected': 'Rechazados',
        }
        for k in status_order:
            if status_counts.get(k, 0) > 0:
                status_labels.append(status_labels_map.get(k, k))
                status_data.append(status_counts[k])

    except Exception as e:
        logger.error(f"Error loading analytics data from DB: {e}")

    # ── CSV Stats ──
    try:
        csv_files = glob.glob(os.path.join('results', '*.csv'))
        for fpath in sorted(csv_files, key=os.path.getmtime, reverse=True)[:15]:
            fname = os.path.basename(fpath)
            try:
                df_tmp = pd.read_csv(fpath, encoding='latin-1', nrows=1)
                count  = sum(1 for _ in open(fpath, encoding='latin-1')) - 1
            except Exception:
                count = 0

            # Detectar plataforma desde nombre de archivo
            platform = 'otro'
            for p in ['indeed','linkedin','computrabajo','bumeran']:
                if p in fname.lower():
                    platform = p
                    break

            mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%d/%m/%Y %H:%M')
            csv_stats.append({'name': fname, 'count': count, 'platform': platform, 'date': mtime})
    except Exception as e:
        logger.error(f"Error loading CSV stats: {e}")

    return render_template(
        'analytics.html',
        stats=stats,
        timeline_labels=timeline_labels,
        timeline_data=timeline_data,
        status_labels=status_labels,
        status_data=status_data,
        csv_stats=csv_stats,
    )


@app.route('/api/tracker/stats')
def tracker_stats():
    """Devuelve conteos por estado para el embudo del Job Tracker."""
    try:
        from collections import defaultdict
        all_jobs = db_manager.get_all_tracked_jobs()
        counts   = defaultdict(int)
        for job in all_jobs:
            counts[job.get('status', 'unknown')] += 1
        return jsonify({'success': True, 'stats': dict(counts)})
    except Exception as e:
        logger.error(f"Error in tracker stats: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# ANALYTICS AVANZADO – endpoint JSON (para llamadas AJAX)
# ─────────────────────────────────────────────────────────────

@app.route('/api/analytics/advanced')
def api_analytics_advanced():
    """Métricas avanzadas: tiempos promedio por etapa, tasa respuesta, heatmap."""
    from collections import defaultdict

    try:
        all_jobs = db_manager.get_all_tracked_jobs()

        # ── Conteos por estado ──
        by_status = defaultdict(int)
        by_weekday = defaultdict(int)  # 0=Lun … 6=Dom
        total = len(all_jobs)

        applied = interviewing = accepted = rejected = 0

        for job in all_jobs:
            st = job.get('status', 'unknown')
            by_status[st] += 1

            # Detección de día de la semana de postulación
            raw_date = job.get('applied_date') or job.get('date_saved')
            if raw_date:
                try:
                    if hasattr(raw_date, 'weekday'):
                        by_weekday[raw_date.weekday()] += 1
                    else:
                        from datetime import datetime as _dt
                        parsed = _dt.strptime(str(raw_date)[:10], '%Y-%m-%d')
                        by_weekday[parsed.weekday()] += 1
                except Exception:
                    pass

            if st == 'applied':     applied += 1
            if st == 'interviewing': interviewing += 1
            if st == 'accepted':    accepted += 1
            if st == 'rejected':    rejected += 1

        # ── Tasas (evitar divisiones por cero) ──
        response_rate   = round(interviewing / applied * 100, 1) if applied  > 0 else 0
        success_rate    = round(accepted / total * 100, 1)       if total    > 0 else 0
        rejection_rate  = round(rejected / total * 100, 1)       if total    > 0 else 0

        # ── Heatmap por día de semana ──
        weekday_labels = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        heatmap = [{'day': weekday_labels[i], 'count': by_weekday.get(i, 0)} for i in range(7)]

        # ── Insights automáticos ──
        insights = []
        if response_rate >= 20:
            insights.append({'icon': '🎯', 'type': 'success', 'text': f'Excelente tasa de respuesta ({response_rate}%). ¡Tus cartas de presentación están funcionando!'})
        elif response_rate > 0:
            insights.append({'icon': '💡', 'type': 'warning', 'text': f'Tasa de respuesta del {response_rate}%. Considera personalizar más tu CV para cada empresa.'})
        else:
            insights.append({'icon': '📝', 'type': 'info', 'text': 'Comienza a postularte para ver métricas de respuesta.'})

        if by_status.get('bookmarked', 0) > by_status.get('applied', 0):
            insights.append({'icon': '🚀', 'type': 'info', 'text': 'Tienes más empleos guardados que postulaciones. ¡Es momento de aplicar!'})

        best_day = max(by_weekday, key=lambda k: by_weekday[k], default=None)
        if best_day is not None and by_weekday[best_day] > 0:
            insights.append({'icon': '📅', 'type': 'info', 'text': f'Tu día más activo es el {weekday_labels[best_day]}.'})

        # ── Embudo detallado con conversiones ──
        funnel = []
        stages = [
            ('bookmarked',  'Guardados'),
            ('applying',    'Aplicando'),
            ('applied',     'Postulados'),
            ('interviewing','Entrevistas'),
            ('negotiating', 'Negociando'),
            ('accepted',    'Aceptados'),
        ]
        prev_count = total if total > 0 else 1
        for key, label in stages:
            count = by_status.get(key, 0)
            conversion = round(count / prev_count * 100, 1) if prev_count > 0 else 0
            funnel.append({'key': key, 'label': label, 'count': count, 'conversion': conversion})
            if count > 0:
                prev_count = count

        return jsonify({
            'success': True,
            'total': total,
            'response_rate': response_rate,
            'success_rate': success_rate,
            'rejection_rate': rejection_rate,
            'by_status': dict(by_status),
            'heatmap': heatmap,
            'funnel': funnel,
            'insights': insights,
        })

    except Exception as e:
        logger.error(f"Error in advanced analytics: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# EXPORTACIÓN – Excel y PDF
# ─────────────────────────────────────────────────────────────

@app.route('/export/tracker/excel')
def export_tracker_excel():
    """Exporta el Job Tracker completo a Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        jobs = db_manager.get_all_tracked_jobs()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Tracker"

        # ── Estilos ──
        header_fill   = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font   = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
        center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align    = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border   = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        status_colors = {
            'bookmarked': "FEF3C7", 'applying': "DBEAFE",
            'applied': "E0E7FF", 'interviewing': "D1FAE5",
            'negotiating': "FDE68A", 'accepted': "A7F3D0", 'rejected': "FEE2E2"
        }

        # ── Encabezados ──
        columns = ['ID', 'Título', 'Empresa', 'Ubicación', 'Salario',
                   'Estado', 'Entusiasmo', 'Fecha Guardado', 'Fecha Aplicado',
                   'Deadline', 'Notas', 'Enlace']
        db_keys = ['id', 'title', 'company', 'location', 'salary',
                   'status', 'excitement', 'date_saved', 'date_applied',
                   'deadline', 'notes', 'job_link']

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 22

        # ── Datos ──
        for row_idx, job in enumerate(jobs, start=2):
            status = job.get('status', '')
            row_fill_color = status_colors.get(status, "FFFFFF")
            row_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")

            for col_idx, key in enumerate(db_keys, start=1):
                value = job.get(key, '')
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = left_align if col_idx > 2 else center_align
            ws.row_dimensions[row_idx].height = 18

        # ── Ancho de columnas ──
        col_widths = [6, 32, 22, 18, 14, 15, 12, 16, 16, 14, 30, 40]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ── "Freeze" primera fila ──
        ws.freeze_panes = 'A2'

        # ── Hoja de resumen ──
        ws2 = wb.create_sheet("Resumen")
        from collections import Counter
        statuses = Counter(j.get('status', 'unknown') for j in jobs)
        ws2.append(['Estado', 'Cantidad', '% del Total'])
        for st, cnt in sorted(statuses.items(), key=lambda x: -x[1]):
            pct = round(cnt / len(jobs) * 100, 1) if jobs else 0
            ws2.append([st, cnt, f"{pct}%"])

        # ── Guardar en memoria ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"EmpleoIA_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Error exporting tracker to Excel: {e}")
        flash(f'Error al exportar: {str(e)}', 'danger')
        return redirect(url_for('tracker'))


@app.route('/export/tracker/pdf')
def export_tracker_pdf():
    """Exporta el Job Tracker a PDF con ReportLab."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        jobs = db_manager.get_all_tracked_jobs()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Title'],
                                     fontSize=18, textColor=colors.HexColor('#667EEA'),
                                     spaceAfter=6)
        subtitle_style = ParagraphStyle('sub', parent=styles['Normal'],
                                        fontSize=10, textColor=colors.HexColor('#6B7280'),
                                        spaceAfter=16)

        elements = []
        elements.append(Paragraph("EmpleoIA – Reporte de Postulaciones", title_style))
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(jobs)} registros totales", subtitle_style))
        elements.append(Spacer(1, 10))

        # ── Tabla de trabajos ──
        col_names = ['Título', 'Empresa', 'Ubicación', 'Estado', 'Fecha Aplicado', 'Salario']
        db_keys   = ['title', 'company', 'location', 'status', 'date_applied', 'salary']
        data = [col_names]
        for job in jobs[:200]:   # límite de 200 filas en PDF
            row = []
            for key in db_keys:
                val = job.get(key, '')
                if hasattr(val, 'strftime'):
                    val = val.strftime('%d/%m/%Y')
                row.append(str(val)[:40] if val else '-')
            data.append(row)

        col_widths_pdf = [7*cm, 6*cm, 5*cm, 3.5*cm, 3.5*cm, 3.5*cm]
        table = Table(data, colWidths=col_widths_pdf, repeatRows=1)

        status_color_map = {
            'bookmarked': colors.HexColor('#FEF3C7'),
            'applying':   colors.HexColor('#DBEAFE'),
            'applied':    colors.HexColor('#E0E7FF'),
            'interviewing': colors.HexColor('#D1FAE5'),
            'negotiating': colors.HexColor('#FDE68A'),
            'accepted':   colors.HexColor('#A7F3D0'),
            'rejected':   colors.HexColor('#FEE2E2'),
        }

        base_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]

        # Color por estado
        for i, job in enumerate(jobs[:200], start=1):
            st = job.get('status', '')
            if st in status_color_map:
                base_style.append(('BACKGROUND', (3, i), (3, i), status_color_map[st]))

        table.setStyle(TableStyle(base_style))
        elements.append(table)

        doc.build(elements)
        output.seek(0)

        filename = f"EmpleoIA_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        logger.error(f"Error exporting tracker to PDF: {e}")
        flash(f'Error al exportar PDF: {str(e)}', 'danger')
        return redirect(url_for('tracker'))


@app.route('/export/results/<filename>/excel')
def export_results_excel(filename):
    """Exporta un CSV de resultados como Excel formateado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        file_path = os.path.join('results', filename)
        if not os.path.exists(file_path) or not filename.endswith('.csv'):
            flash('Archivo no encontrado', 'danger')
            return redirect(url_for('results'))

        df = pd.read_csv(file_path, encoding='latin-1')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"

        header_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border  = Border(
            left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
        )

        # Encabezados
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 20

        # Datos
        alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            row_fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            ws.row_dimensions[row_idx].height = 15

        # Auto-width
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        base = filename.replace('.csv', '')
        dl_name = f"{base}_export.xlsx"
        return send_file(output, as_attachment=True, download_name=dl_name,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Error exporting results to Excel: {e}")
        flash(f'Error al exportar: {str(e)}', 'danger')
        return redirect(url_for('results'))


# ─────────────────────────────────────────────
# Error Handlers
# ─────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, message='Página no encontrada'), 404

@app.errorhandler(500)
def server_error(e):
    logger.error(f"500 Internal Server Error: {e}")
    return render_template('error.html', code=500, message=str(e)), 500


# ─────────────────────────────────────────────
# SocketIO: progreso del scraper en tiempo real
# ─────────────────────────────────────────────

@socketio.on('connect')
def on_connect():
    emit('connected', {'data': 'Connected to EmpleoIA SocketIO'})

@socketio.on('check_status')
def handle_check_status(data):
    """Cliente puede pedir el estado actual de cualquier scraper."""
    scraper_key = data.get('scraper', 'indeed_scraping')
    status = job_status.get(scraper_key, {})
    emit('scraper_status', status)


def emit_scraper_progress(scraper_key: str):
    """Emite el progreso del scraper via SocketIO (llamar desde los threads de scraping)."""
    try:
        socketio.emit('scraper_status', {
            'scraper': scraper_key,
            **job_status.get(scraper_key, {})
        })
    except Exception:
        pass  # No romper el scraper si SocketIO falla


if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', debug=True, allow_unsafe_werkzeug=True)
