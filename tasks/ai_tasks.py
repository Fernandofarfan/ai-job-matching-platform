"""
tasks/ai_tasks.py — Tareas de IA asíncronas vía Celery
Para procesamiento por lotes, análisis semántico masivo y exportaciones.
"""
import os
import glob
import logging
from datetime import datetime

from celery_app import celery

logger = logging.getLogger(__name__)


@celery.task(bind=True, name="tasks.ai_tasks.batch_analyze_jobs")
def batch_analyze_jobs(self, csv_path: str, resume_text: str = ""):
    """Analiza semánticamente todas las ofertas de un CSV con IA."""
    import pandas as pd
    from ai_engine import analyze_job_complete

    try:
        self.update_state(state="PROGRESS", meta={"progress": 5, "message": "Cargando CSV..."})
        df = pd.read_csv(csv_path, encoding="latin-1")
        total = len(df)

        results = []
        desc_col = next((c for c in ["Description", "Job_Description", "description"] if c in df.columns), None)
        title_col = next((c for c in ["Title", "title", "Titulo"] if c in df.columns), None)

        for i, row in df.iterrows():
            progress = 10 + int((i / max(total, 1)) * 85)
            self.update_state(state="PROGRESS", meta={
                "progress": progress,
                "message": f"Analizando oferta {i+1}/{total}..."
            })

            title = str(row.get(title_col, "")) if title_col else "Sin título"
            description = str(row.get(desc_col, "")) if desc_col else ""

            if not description or len(description) < 50:
                results.append({"ai_match_score": 0, "ai_status": "sin_descripcion"})
                continue

            try:
                analysis = analyze_job_complete(title, description, resume_text)
                results.append({
                    "ai_match_score": analysis.get("match_score", 0),
                    "ai_nivel": analysis.get("nivel_experiencia", ""),
                    "ai_modalidad": analysis.get("modalidad", ""),
                    "ai_skills_faltantes": ", ".join(analysis.get("skills_faltantes", [])[:5]),
                    "ai_resumen": analysis.get("resumen_ejecutivo", ""),
                    "ai_status": "ok"
                })
            except Exception as e:
                results.append({"ai_match_score": 0, "ai_status": f"error: {str(e)}"})

        # Agregar columnas de IA al DataFrame
        result_df = df.copy()
        for key in results[0].keys():
            result_df[key] = [r.get(key, "") for r in results]

        # Guardar versión enriquecida
        output_path = csv_path.replace(".csv", "_ai_analyzed.csv")
        result_df.to_csv(output_path, index=False)

        return {
            "status": "completed",
            "analyzed": total,
            "output_file": os.path.basename(output_path),
            "message": f"Análisis IA completado: {total} ofertas procesadas"
        }

    except Exception as e:
        logger.error(f"Batch AI analysis failed: {e}", exc_info=True)
        raise


@celery.task(bind=True, name="tasks.ai_tasks.batch_generate_cover_letters")
def batch_generate_cover_letters(self, csv_path: str, resume_text: str, tone: str = "semi-formal"):
    """Genera cartas de presentación para todas las ofertas de un CSV."""
    import pandas as pd

    try:
        self.update_state(state="PROGRESS", meta={"progress": 5, "message": "Iniciando generación masiva..."})

        df = pd.read_csv(csv_path, encoding="latin-1")
        total = len(df)
        output_dir = os.path.join("temp", "cover_letters_batch")
        os.makedirs(output_dir, exist_ok=True)

        desc_col = next((c for c in ["Description", "Job_Description"] if c in df.columns), None)
        title_col = next((c for c in ["Title", "title"] if c in df.columns), None)
        company_col = next((c for c in ["Company", "company", "Empresa"] if c in df.columns), None)

        generated = 0
        for i, row in df.iterrows():
            progress = 10 + int((i / max(total, 1)) * 85)
            self.update_state(state="PROGRESS", meta={
                "progress": progress,
                "message": f"Generando carta {i+1}/{total}..."
            })

            title = str(row.get(title_col, "Posicion")) if title_col else "Posicion"
            company = str(row.get(company_col, "Empresa")) if company_col else "Empresa"
            description = str(row.get(desc_col, "")) if desc_col else ""

            if not description:
                continue

            try:
                # Usar función de generación de cartas existente
                from ai_engine import _call_gemini
                prompt = f"""Generá una carta de presentación {tone} en español rioplatense para este puesto.
Empresa: {company}
Puesto: {title}
Descripción: {description[:1500]}
CV/Perfil del candidato: {resume_text[:1000]}

La carta debe tener 3 párrafos: apertura impactante, experiencia relevante, cierre con call-to-action.
Máximo 300 palabras."""
                letter = _call_gemini(prompt)

                filename = f"carta_{i+1:03d}_{company[:20].replace(' ', '_')}.txt"
                with open(os.path.join(output_dir, filename), "w", encoding="utf-8") as f:
                    f.write(f"Empresa: {company}\nPuesto: {title}\n\n{letter}")
                generated += 1

            except Exception as e:
                logger.error(f"Error generating cover letter for row {i}: {e}")

        return {
            "status": "completed",
            "generated": generated,
            "total": total,
            "output_dir": output_dir,
            "message": f"Generadas {generated}/{total} cartas de presentación"
        }

    except Exception as e:
        logger.error(f"Batch cover letter generation failed: {e}", exc_info=True)
        raise


@celery.task(name="tasks.ai_tasks.cleanup_old_temp_files")
def cleanup_old_temp_files(days: int = 7):
    """Limpia archivos temporales más viejos que N días."""
    import time as time_mod

    cleaned = 0
    cutoff = time_mod.time() - (days * 86400)

    dirs_to_clean = [
        os.path.join("temp", "test_tasks"),
        os.path.join("temp", "resumes"),
    ]

    for dir_path in dirs_to_clean:
        if not os.path.exists(dir_path):
            continue
        for filepath in glob.glob(os.path.join(dir_path, "*")):
            try:
                if os.path.getmtime(filepath) < cutoff:
                    os.remove(filepath)
                    cleaned += 1
            except Exception as e:
                logger.warning(f"Could not delete {filepath}: {e}")

    logger.info(f"Cleanup: removed {cleaned} old temp files")
    return {"cleaned": cleaned, "cutoff_days": days}


@celery.task(bind=True, name="tasks.ai_tasks.export_tracker_to_excel")
def export_tracker_to_excel(self, job_list: list, output_path: str):
    """Exporta el Job Tracker a Excel con formato premium."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        self.update_state(state="PROGRESS", meta={"progress": 10, "message": "Creando Excel..."})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Tracker EmpleoIA"

        # Estilos
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        status_colors = {
            "bookmarked": "E3F2FD",
            "applying":   "FFF9C4",
            "applied":    "F3E5F5",
            "interviewing": "E8F5E9",
            "negotiating":  "FFF3E0",
            "accepted":     "C8E6C9",
            "rejected":     "FFCDD2",
        }

        headers = ["#", "Título", "Empresa", "Ubicación", "Salario", "Estado",
                   "Fecha Aplicación", "Follow-up", "Notas", "Enlace"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 25

        for row_idx, job in enumerate(job_list, 2):
            status = job.get("status", "bookmarked")
            row_fill = PatternFill(
                start_color=status_colors.get(status, "FFFFFF"),
                end_color=status_colors.get(status, "FFFFFF"),
                fill_type="solid"
            )
            data = [
                row_idx - 1,
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("salary", ""),
                status,
                str(job.get("applied_date", ""))[:10],
                str(job.get("follow_up", ""))[:10] if job.get("follow_up") else "",
                job.get("notes", ""),
                job.get("job_link", ""),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        self.update_state(state="PROGRESS", meta={"progress": 100, "message": "Excel listo"})

        return {"status": "completed", "output_path": output_path}

    except Exception as e:
        logger.error(f"Excel export task failed: {e}", exc_info=True)
        raise
